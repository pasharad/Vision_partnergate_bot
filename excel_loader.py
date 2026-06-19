import logging
from datetime import datetime

import openpyxl

import config
from models import Board, DataCache

logger = logging.getLogger(__name__)

cache = DataCache()


def load() -> bool:
    try:
        wb = openpyxl.load_workbook(config.EXCEL_FILE_PATH, data_only=True)
    except FileNotFoundError:
        logger.error(f"Excel file not found: {config.EXCEL_FILE_PATH}")
        return False
    except Exception as e:
        logger.error(f"Failed to open Excel file: {e}")
        return False

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    boards = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        board_code = row[3]  # D
        if not board_code or board_code == "خدمات اضافی":
            continue

        try:
            board = Board(
                board_code=str(board_code),
                station=str(row[4] or ""),       # E
                status=str(row[5] or ""),         # F
                reservation_end=str(row[6] or ""),# G
                city=str(row[7] or ""),           # H
                description=str(row[8] or ""),    # I
                mode=str(row[9] or ""),           # J
                price_monthly=int(row[10] or 0),  # K
                price_2month=int(row[11] or 0),   # L
                price_3_5month=int(row[12] or 0), # M
                price_6_8month=int(row[13] or 0), # N
                area_sqm=float(row[14] or 0),     # O
                print_unit_price=int(row[15] or 0),# P
                print_total_cost=int(row[16] or 0),# Q
            )
            boards.append(board)
        except Exception as e:
            logger.warning(f"Skipping row {row_idx}: {e}")
            continue

    cache.boards = boards
    cache.last_loaded = datetime.now()
    logger.info(f"Loaded {len(boards)} boards from Excel")
    return True


def reload() -> bool:
    return load()
